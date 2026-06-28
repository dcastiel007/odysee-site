#!/usr/bin/env python3
"""
build-views.py — regenerate internal/tracker.html, internal/gantt.html, and
the PUBLIC pages/project-plan.html from ticoprojectplanv2.xlsx (single source).

Usage:
    python build-views.py [plan.xlsx] [out_dir]
      plan.xlsx  default: ticoprojectplanv2.xlsx (same folder)
      out_dir    default: internal/

internal/tracker.html, internal/gantt.html  — self-contained Brand pages (RTL).
pages/project-plan.html                      — PUBLIC page: clones the live V5
    chrome from pages/agents.html (head/header/footer) and injects a gantt +
    progress summary. Always inherits whatever site chrome is current, so it
    never drifts from the rest of the site. Generated — do not hand-edit.

Re-run after every xlsx edit, then git add + commit + push.
"""
import sys, re
import datetime as dt
from pathlib import Path
import html as _html
try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed. Run: pip install openpyxl --break-system-packages")

STATUS = {
    "KEY":      ("#E5763C", "אבן דרך"),
    "DESIGN":   ("#F2C14E", "עיצוב"),
    "CONTENT":  ("#E5763C", "תוכן"),
    "DEV":      ("#2F8F93", "פיתוח"),
    "IMPL":     ("#1C6B6F", "יישום"),
    "QA":       ("#E5A13C", "בדיקות"),
    "FIX":      ("#D2553F", "תיקונים"),
    "DOCS":     ("#3A5D83", "תיעוד"),
    "PRINT":    ("#1D3D60", "הדפסה"),
    "HW":       ("#6B7588", "חומרה"),
    "SUBMIT":   ("#E5763C", "הגשה"),
    "SETUP":    ("#3A5D83", "הקמה"),
    "LIVE":     ("#E5763C", "תערוכה"),
    "TEARDOWN": ("#6B7588", "פירוק"),
}
DEFAULT_COLOR = "#6B7588"
E = lambda s: _html.escape(str(s)) if s is not None else ""
def hex_to_rgba(hexcol, alpha):
    h = hexcol.lstrip("#"); r,g,b = int(h[0:2],16),int(h[2:4],16),int(h[4:6],16)
    return f"rgba({r},{g},{b},{alpha})"

# ── Read the workbook ──
def read_plan(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Project Plan"]
    phases = []; current = None
    for r in range(5, ws.max_row + 1):
        a = ws.cell(r, 1).value; phase = ws.cell(r, 2).value; task = ws.cell(r, 3).value
        if a is None and task is None: continue
        if phase is None and isinstance(a, str):
            current = {"name": a.strip(), "tasks": []}; phases.append(current); continue
        if current is None:
            current = {"name": "", "tasks": []}; phases.append(current)
        done = isinstance(task, str) and task.strip().startswith("✅")
        current["tasks"].append({
            "num": a, "phase": phase, "task": task,
            "start": ws.cell(r, 4).value, "end": ws.cell(r, 5).value,
            "days": ws.cell(r, 6).value, "status": (ws.cell(r, 7).value or "").strip(),
            "notes": ws.cell(r, 8).value, "done": done,
        })
    title = (ws.cell(1, 1).value or "Odysee — Project Plan")
    title = title.replace("t.Co.", "Odysee").replace("t.Co", "Odysee")
    meta = (ws.cell(2, 1).value or "")
    return phases, title, meta

def all_tasks(phases):
    return [t for p in phases for t in p["tasks"]]

# ===================== INTERNAL pages (unchanged) =====================
def head(title, active):
    nav = [("../index.html","Home",False),("tracker.html","Tracker",active=="tracker"),
           ("gantt.html","Gantt",active=="gantt")]
    navhtml = "\n".join(
        f'      <li><a href="{u}"{" class=\"active\"" if act else ""}>{n}</a></li>'
        for u,n,act in nav)
    roundel = ('<svg class="bm__o" viewBox="0 0 1000 1000" aria-hidden="true" fill="none">'
        '<path class="ring" fill="none" stroke-width="50" stroke-linecap="round" stroke-linejoin="round" d="M 941.072449 422.659241 C 932.16748 320.875366 831.286194 246.557617 715.747742 256.665955 C 600.20929 266.774231 513.765747 357.480774 522.670654 459.264648 C 531.575623 561.048523 632.456909 635.366272 747.995361 625.257935 C 863.533813 615.149658 949.977356 524.443115 941.072449 422.659241 Z"/>'
        '<path class="ring" fill="none" stroke-width="50" stroke-linecap="round" stroke-linejoin="round" d="M 478.329315 459.264648 C 487.234253 357.480774 400.79071 266.774231 285.252258 256.665955 C 169.713791 246.557617 68.832489 320.875366 59.927555 422.659241 C 51.022621 524.443115 137.466171 615.149658 253.004623 625.257935 C 368.543091 635.366272 469.424408 561.048523 478.329315 459.264648 Z"/>'
        '<path class="ring" fill="none" stroke-width="50" stroke-linecap="round" stroke-linejoin="round" d="M 900 500 C 900 279.086121 720.913879 100 500 100 C 279.086121 100 100 279.086121 100 500 C 100 720.913879 279.086121 900 500 900 C 720.913879 900 900 720.913879 900 500 Z"/>'
        '<circle class="cap-end" cx="725" cy="399" r="48"/>'
        '<circle class="cap-end" cx="300" cy="399" r="48"/></svg>')
    return f"""<!DOCTYPE html>
<html lang="he" dir="rtl" data-theme="system">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{E(title)}</title>
<meta name="robots" content="noindex">
<link rel="stylesheet" href="../brand/odysee-tokens.css">
<link rel="stylesheet" href="../assets/css/site.css">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Sora:wght@400;500;700;800&family=Heebo:wght@400;500;700;800&family=JetBrains+Mono:wght@400;500&display=swap">
<style>
{COMPONENT_CSS}
</style>
</head>
<body>
<div class="theme-bar">
  <button class="theme-btn" onclick="setTheme('light')" id="btn-light">☀ בהיר</button>
  <button class="theme-btn" onclick="setTheme('system')" id="btn-system">⊙ מערכת</button>
  <button class="theme-btn" onclick="setTheme('dark')" id="btn-dark">● כהה</button>
</div>
<header class="site-header" dir="ltr">
  <div class="iv-brand">
    <a href="../index.html" class="bm" aria-label="Odysee — home">
      {roundel}<span class="bm__text">dysee</span>
    </a>
    <span class="iv-tagline">Internal · {E(active.title())}</span>
  </div>
  <nav><ul class="site-nav">
{navhtml}
  </ul></nav>
</header>
"""

FOOT = """<footer class="site-footer">
  <div class="site-footer__left">David Castiel · M.Design · HIT 2026 — Internal view, generated from xlsx</div>
  <div class="site-footer__right" dir="ltr">Generated %s</div>
</footer>
<script>
const root=document.documentElement,mq=matchMedia('(prefers-color-scheme: dark)');
let cur=localStorage.getItem('odysee-theme')||'dark';
function apply(t){root.setAttribute('data-theme',t==='system'?(mq.matches?'dark':'light'):t);
  ['light','system','dark'].forEach(x=>document.getElementById('btn-'+x).classList.toggle('active',x===t));
  cur=t;localStorage.setItem('odysee-theme',t);}
function setTheme(t){apply(t);}
mq.addEventListener('change',()=>{if(cur==='system')apply('system');});
apply(cur);
</script>
</body>
</html>"""

COMPONENT_CSS = """
.iv-brand{display:flex;flex-direction:column;gap:4px}
.bm{display:inline-flex;align-items:center;gap:2px;text-decoration:none;line-height:.9}
.bm__o{width:2.2em;height:2.2em}
.bm__text{font-family:var(--od-font-sans,'Sora',sans-serif);font-weight:800;font-size:34px;letter-spacing:-.035em;color:var(--od-fg)}
.bm__o .ring{stroke:var(--od-navy)}.bm__o .cap-end{fill:var(--od-red)}
[data-theme="dark"] .bm__o .ring{stroke:var(--od-paper)}
.iv-tagline{font-size:10px;letter-spacing:.22em;text-transform:uppercase;color:var(--od-fg-subtle);font-weight:500;margin-top:2px}
.iv-wrap{max-width:1200px;margin:0 auto;padding:32px 24px 64px}
.iv-h1{font-family:var(--od-font-sans,'Sora',sans-serif);font-size:28px;font-weight:800;letter-spacing:-.02em;color:var(--od-fg);margin-bottom:4px}
.iv-meta{font-size:12px;color:var(--od-fg-subtle);margin-bottom:8px;direction:ltr;text-align:left}
.iv-summary{display:flex;gap:10px;flex-wrap:wrap;margin:18px 0 28px}
.iv-stat{background:var(--od-bg-sunken);border:1px solid var(--od-border);border-radius:var(--od-r-md,10px);padding:12px 18px;min-width:96px}
.iv-stat .n{font-family:var(--od-font-sans,'Sora',sans-serif);font-size:24px;font-weight:800;color:var(--od-fg)}
.iv-stat .l{font-size:11px;color:var(--od-fg-subtle);letter-spacing:.04em}
.iv-phase{margin-top:30px}
.iv-phase h2{font-family:var(--od-font-sans,'Sora',sans-serif);font-size:16px;font-weight:700;color:var(--od-fg);margin-bottom:2px}
.iv-phase .sub{font-size:11px;color:var(--od-fg-subtle);direction:ltr;text-align:right;margin-bottom:12px}
.pill{display:inline-block;padding:2px 10px;border-radius:999px;font-size:9px;font-weight:600;letter-spacing:.10em;text-transform:uppercase;font-family:var(--od-font-sans,'Sora',sans-serif);white-space:nowrap;direction:ltr}
.iv-table{width:100%;border-collapse:collapse;font-size:13px;font-family:var(--od-font-sans,'Sora',sans-serif)}
.iv-table th{background:var(--od-navy);color:#fff;font-size:9px;font-weight:600;letter-spacing:.10em;text-transform:uppercase;padding:9px 12px;text-align:right}
.iv-table td{padding:9px 12px;color:var(--od-fg-muted);border-bottom:1px solid var(--od-border);vertical-align:top}
.iv-table tr:nth-child(even) td{background:var(--od-bg-sunken)}
.iv-table .c-num{font-family:var(--od-font-mono,monospace);font-size:11px;color:var(--od-fg-subtle);direction:ltr}
.iv-table .c-task{color:var(--od-fg);font-weight:500}
.iv-table .c-date{font-family:var(--od-font-mono,monospace);font-size:11px;color:var(--od-fg-subtle);direction:ltr;white-space:nowrap}
.iv-table .c-notes{font-size:12px;color:var(--od-fg-subtle);line-height:1.5}
.gantt{margin-top:8px;font-family:var(--od-font-sans,'Sora',sans-serif);direction:ltr}
.gantt-months{display:flex;border-bottom:1px solid var(--od-border-strong);margin-left:230px;position:relative;height:22px}
.gantt-month{font-size:10px;color:var(--od-fg-subtle);letter-spacing:.06em;border-right:1px solid var(--od-border);padding:2px 0 0 4px;box-sizing:border-box;text-align:left}
.gantt-row{display:flex;align-items:center;height:30px;border-bottom:1px solid var(--od-border)}
.gantt-label{width:230px;flex-shrink:0;font-size:12px;color:var(--od-fg);padding:0 10px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;direction:rtl;text-align:right;display:flex;align-items:center;gap:8px}
.gantt-label .gtext{overflow:hidden;text-overflow:ellipsis;flex:1}
.gantt-label .gnum{font-family:var(--od-font-mono,monospace);font-size:10px;color:var(--od-fg-subtle);direction:ltr;flex-shrink:0;min-width:22px;text-align:left;order:2}
.gantt-track{position:relative;flex:1;height:100%}
.gantt-bar{position:absolute;top:7px;height:16px;border-radius:4px;opacity:.9;display:flex;align-items:center;direction:ltr}
.gantt-phase{font-family:var(--od-font-sans,'Sora',sans-serif);font-size:11px;font-weight:700;color:var(--od-red);letter-spacing:.06em;text-transform:uppercase;padding:14px 10px 4px;border-bottom:1px solid var(--od-border-strong);margin-top:6px;direction:rtl;text-align:right}
.gantt-today{position:absolute;top:0;bottom:0;width:2px;background:var(--od-red);z-index:5}
.gantt-legend{display:flex;gap:14px;flex-wrap:wrap;margin-top:22px;padding-top:16px;border-top:1px solid var(--od-border);direction:rtl}
.gantt-legend .lg{display:flex;align-items:center;gap:6px;font-size:11px;color:var(--od-fg-muted)}
.gantt-legend .sw{width:12px;height:12px;border-radius:3px}
@media(max-width:640px){.gantt-months{margin-left:130px}.gantt-label{width:130px;font-size:11px}.iv-table .c-notes{display:none}}
"""

def pill(status):
    color,_ = STATUS.get(status,(DEFAULT_COLOR,status)); bg = hex_to_rgba(color,0.16)
    return f'<span class="pill" style="background:{bg};color:{color}">{E(status)}</span>'
def fmt(d):
    return d.strftime("%Y-%m-%d") if isinstance(d,dt.datetime) else E(d)

def build_tracker(phases, title, meta):
    tasks = all_tasks(phases)
    from collections import Counter
    counts = Counter(t["status"] for t in tasks)
    stats = "".join(f'<div class="iv-stat"><div class="n">{len(tasks)}</div><div class="l">משימות</div></div>' if i==0 else "" for i in range(1))
    stats += f'<div class="iv-stat"><div class="n">{len(phases)}</div><div class="l">פאזות</div></div>'
    for st,c in counts.most_common(4):
        col = STATUS.get(st,(DEFAULT_COLOR,))[0]
        stats += (f'<div class="iv-stat"><div class="n" style="color:{col}">{c}</div><div class="l">{E(st)}</div></div>')
    body = [f'<div class="iv-wrap">',f'<h1 class="iv-h1">מעקב פרויקט</h1>',f'<div class="iv-meta">{E(meta)}</div>',f'<div class="iv-summary">{stats}</div>']
    for p in phases:
        if not p["tasks"]: continue
        body.append(f'<div class="iv-phase"><h2>{E(p["name"])}</h2>')
        body.append('<table class="iv-table"><thead><tr><th style="width:36px">#</th><th>משימה</th><th style="width:90px">התחלה</th><th style="width:90px">סיום</th><th style="width:50px">ימים</th><th style="width:90px">סטטוס</th><th>הערות</th></tr></thead><tbody>')
        for t in p["tasks"]:
            body.append(f'<tr><td class="c-num">{E(t["num"])}</td><td class="c-task">{E(t["task"])}</td><td class="c-date">{fmt(t["start"])}</td><td class="c-date">{fmt(t["end"])}</td><td class="c-num">{E(t["days"])}</td><td>{pill(t["status"])}</td><td class="c-notes">{E(t["notes"])}</td></tr>')
        body.append('</tbody></table></div>')
    body.append('</div>')
    return head("Odysee — מעקב פרויקט","tracker") + "\n".join(body) + (FOOT % dt.date.today().isoformat())

def build_gantt(phases, title, meta):
    tasks = all_tasks(phases)
    starts = [t["start"] for t in tasks if isinstance(t["start"],dt.datetime)]
    ends = [t["end"] for t in tasks if isinstance(t["end"],dt.datetime)]
    p0,p1 = min(starts),max(ends); total = (p1-p0).days or 1
    def pct(d): return (d-p0).days/total*100
    months=[]; cur=dt.datetime(p0.year,p0.month,1)
    while cur<=p1:
        nxt=dt.datetime(cur.year+(cur.month==12),(cur.month%12)+1,1)
        seg_start=max(cur,p0); seg_end=min(nxt,p1); w=(seg_end-seg_start).days/total*100
        months.append((cur.strftime("%b %Y"),w)); cur=nxt
    months_html="".join(f'<div class="gantt-month" style="width:{w:.3f}%">{m}</div>' for m,w in months)
    today=dt.datetime.now(); today_marker=""
    if p0<=today<=p1:
        today_marker=(f'<div class="gantt-today" style="left:{pct(today):.3f}%" title="today {today.date()}"></div>')
    rows=[]
    for p in phases:
        if not p["tasks"]: continue
        rows.append(f'<div class="gantt-phase">{E(p["name"])}</div>')
        for t in p["tasks"]:
            if not (isinstance(t["start"],dt.datetime) and isinstance(t["end"],dt.datetime)): continue
            color=STATUS.get(t["status"],(DEFAULT_COLOR,))[0]; left=pct(t["start"])
            width=max((t["end"]-t["start"]).days/total*100,0.8); label=E(t["task"])
            rows.append(f'<div class="gantt-row"><div class="gantt-label"><span class="gnum">{E(t["num"])}</span><span class="gtext">{label}</span></div><div class="gantt-track"><div class="gantt-bar" style="left:{left:.3f}%;width:{width:.3f}%;background:{color}" title="{label} · {fmt(t["start"])}→{fmt(t["end"])}"></div></div></div>')
    seen=[]
    for t in tasks:
        if t["status"] not in seen: seen.append(t["status"])
    legend="".join(f'<div class="lg"><span class="sw" style="background:{STATUS.get(s,(DEFAULT_COLOR,))[0]}"></span>{E(s)}</div>' for s in seen)
    body=['<div class="iv-wrap">','<h1 class="iv-h1">Gantt — ציר זמן</h1>',f'<div class="iv-meta">{E(meta)}</div>',f'<div class="iv-meta" dir="ltr">{p0.date()} → {p1.date()} · {total} days</div>','<div class="gantt">',f'<div class="gantt-months">{months_html}{today_marker}</div>',"".join(rows),f'<div class="gantt-legend">{legend}</div>','</div></div>']
    return head("Odysee — Gantt","gantt") + "\n".join(body) + (FOOT % dt.date.today().isoformat())

# ===================== PUBLIC page (project-plan.html) =====================
PLAN_CSS = """<style>
.pp-wrap{max-width:1200px;margin:0 auto;padding:8px 24px 80px;direction:rtl}
.pp-kicker{font-family:var(--od-font-sans,'Sora',sans-serif);font-size:11px;font-weight:600;letter-spacing:.18em;text-transform:uppercase;color:var(--od-red);direction:ltr;text-align:right}
.pp-h1{font-family:var(--od-font-sans,'Sora',sans-serif);font-size:34px;font-weight:800;letter-spacing:-.02em;color:var(--od-fg);margin:6px 0 4px}
.pp-meta{font-size:13px;color:var(--od-fg-muted);margin-bottom:20px;line-height:1.6}
.pp-stats{display:flex;gap:12px;flex-wrap:wrap;margin:0 0 30px}
.pp-stat{flex:1;min-width:150px;background:var(--od-bg-sunken,rgba(0,0,0,.04));border:1px solid var(--od-border);border-radius:var(--od-r-lg,16px);padding:18px 22px}
.pp-stat .n{font-family:var(--od-font-sans,'Sora',sans-serif);font-size:32px;font-weight:800;color:var(--od-red);line-height:1}
.pp-stat .l{font-size:12px;color:var(--od-fg-muted);margin-top:6px}
.pp-prog{height:7px;border-radius:999px;background:var(--od-border);margin-top:12px;overflow:hidden}
.pp-prog>span{display:block;height:100%;background:var(--od-red);border-radius:999px}
.pp-section-label{font-family:var(--od-font-sans,'Sora',sans-serif);font-size:11px;font-weight:600;letter-spacing:.16em;text-transform:uppercase;color:var(--od-fg-muted);direction:ltr;text-align:right;margin:8px 0 14px}
.pp-g{font-family:var(--od-font-sans,'Sora',sans-serif);direction:ltr;border:1px solid var(--od-border);border-radius:var(--od-r-lg,16px);overflow:hidden}
.pp-g-months{display:flex;border-bottom:1px solid var(--od-border-strong,var(--od-border));margin-left:240px;position:relative;height:24px;background:var(--od-bg-sunken,transparent)}
.pp-g-month{font-size:10px;color:var(--od-fg-muted);letter-spacing:.05em;border-right:1px solid var(--od-border);padding:4px 0 0 6px;box-sizing:border-box;text-align:left}
.pp-g-row{display:flex;align-items:center;height:32px;border-bottom:1px solid var(--od-border)}
.pp-g-row:last-child{border-bottom:none}
.pp-g-label{width:240px;flex-shrink:0;font-size:12px;color:var(--od-fg);padding:0 12px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;direction:rtl;text-align:right;display:flex;align-items:center;gap:8px}
.pp-g-label .t{overflow:hidden;text-overflow:ellipsis;flex:1}
.pp-g-label .nu{font-family:var(--od-font-mono,monospace);font-size:10px;color:var(--od-fg-subtle);direction:ltr;flex-shrink:0;min-width:22px;text-align:left;order:2}
.pp-g-track{position:relative;flex:1;height:100%}
.pp-g-bar{position:absolute;top:8px;height:16px;border-radius:5px;opacity:.92}
.pp-g-bar.done{opacity:.4}
.pp-g-phase{font-family:var(--od-font-sans,'Sora',sans-serif);font-size:11px;font-weight:700;color:var(--od-red);letter-spacing:.05em;text-transform:uppercase;padding:14px 12px 6px;border-bottom:1px solid var(--od-border-strong,var(--od-border));background:var(--od-bg-sunken,transparent);direction:rtl;text-align:right}
.pp-g-today{position:absolute;top:0;bottom:0;width:2px;background:var(--od-red);z-index:5}
.pp-legend{display:flex;gap:16px;flex-wrap:wrap;margin-top:20px;direction:rtl}
.pp-legend .lg{display:flex;align-items:center;gap:6px;font-size:11px;color:var(--od-fg-muted)}
.pp-legend .sw{width:12px;height:12px;border-radius:3px}
@media(max-width:640px){.pp-g-months{margin-left:130px}.pp-g-label{width:130px;font-size:11px}.pp-h1{font-size:26px}}
</style>"""

def build_public_plan(phases, meta, template_path):
    tpl = template_path.read_text(encoding="utf-8")
    tasks = all_tasks(phases)
    dated = [t for t in tasks if isinstance(t["start"],dt.datetime) and isinstance(t["end"],dt.datetime)]
    p0,p1 = min(t["start"] for t in dated), max(t["end"] for t in dated)
    total=(p1-p0).days or 1
    pct=lambda d:(d-p0).days/total*100
    done_n=sum(1 for t in tasks if t["done"]); tot=len(tasks)
    prog=round(done_n/tot*100) if tot else 0
    # days to submission (from SUBMIT task; fallback 2026-08-04)
    sub=[t["start"] for t in tasks if t["status"]=="SUBMIT" and isinstance(t["start"],dt.datetime)]
    sub_date=(min(sub) if sub else dt.datetime(2026,8,4)).date()
    days_left=(sub_date-dt.date.today()).days
    ntracks=len([p for p in phases if p["tasks"]])
    # months
    months=[]; cur=dt.datetime(p0.year,p0.month,1)
    while cur<=p1:
        nxt=dt.datetime(cur.year+(cur.month==12),(cur.month%12)+1,1)
        w=(min(nxt,p1)-max(cur,p0)).days/total*100; months.append((cur.strftime("%b"),w)); cur=nxt
    mhtml="".join(f'<div class="pp-g-month" style="width:{w:.3f}%">{m}</div>' for m,w in months)
    today=dt.datetime.now(); tmark=""
    if p0<=today<=p1: tmark=f'<div class="pp-g-today" style="left:{pct(today):.3f}%"></div>'
    rows=[]
    for p in phases:
        if not p["tasks"]: continue
        rows.append(f'<div class="pp-g-phase">{E(p["name"])}</div>')
        for t in p["tasks"]:
            if not (isinstance(t["start"],dt.datetime) and isinstance(t["end"],dt.datetime)): continue
            color=STATUS.get(t["status"],(DEFAULT_COLOR,))[0]; left=pct(t["start"])
            width=max((t["end"]-t["start"]).days/total*100,0.8)
            label=E(re.sub(r"^✅\s*","",str(t["task"])))
            dn=" done" if t["done"] else ""
            tip=("✓ " if t["done"] else "")+f'{label} · {fmt(t["start"])}→{fmt(t["end"])}'
            rows.append(f'<div class="pp-g-row"><div class="pp-g-label"><span class="nu">{E(t["num"])}</span><span class="t">{("✓ " if t["done"] else "")+label}</span></div><div class="pp-g-track"><div class="pp-g-bar{dn}" style="left:{left:.3f}%;width:{width:.3f}%;background:{color}" title="{tip}"></div></div></div>')
    seen=[]
    for t in tasks:
        if t["status"] not in seen: seen.append(t["status"])
    legend="".join(f'<div class="lg"><span class="sw" style="background:{STATUS.get(s,(DEFAULT_COLOR,))[0]}"></span>{E(STATUS.get(s,(0,s))[1])}</div>' for s in seen)
    body=f"""
<main class="pp-wrap">
  <div class="pp-kicker">Project Plan · Live</div>
  <h1 class="pp-h1">תכנית עבודה</h1>
  <div class="pp-meta">{E(meta)}</div>
  <div class="pp-stats">
    <div class="pp-stat"><div class="n">{prog}%</div><div class="l">בוצע ({done_n}/{tot})</div><div class="pp-prog"><span style="width:{prog}%"></span></div></div>
    <div class="pp-stat"><div class="n">{days_left}</div><div class="l">ימים להגשה · {sub_date.strftime('%d/%m')}</div></div>
    <div class="pp-stat"><div class="n">{tot}</div><div class="l">משימות</div></div>
    <div class="pp-stat"><div class="n">{ntracks}</div><div class="l">נתיבים</div></div>
  </div>
  <div class="pp-section-label">Gantt · {p0.strftime('%d/%m')} → {p1.strftime('%d/%m/%Y')}</div>
  <div class="pp-g">
    <div class="pp-g-months">{mhtml}{tmark}</div>
    {''.join(rows)}
  </div>
  <div class="pp-legend">{legend}</div>
</main>
"""
    # set title, drop nav active, custom footer-right, inject css before </head>
    tpl=re.sub(r"<title>.*?</title>", "<title>Odysee — Project Plan</title>", tpl, count=1, flags=re.S)
    tpl=tpl.replace(' class="active"','',1)
    tpl=re.sub(r'(<div class="site-footer__right">).*?(</div>)', r'\1Odysee · Project Plan · '+dt.date.today().isoformat()+r'\2', tpl, count=1, flags=re.S)
    tpl=tpl.replace("</head>", PLAN_CSS+"\n</head>", 1)
    i=tpl.index("</header>")+len("</header>")
    j=tpl.index("<footer", i)
    return tpl[:i]+body+tpl[j:]

# ── main ──
def main():
    xlsx=Path(sys.argv[1]) if len(sys.argv)>1 else Path("ticoprojectplanv2.xlsx")
    out=Path(sys.argv[2]) if len(sys.argv)>2 else Path("internal")
    if not xlsx.exists(): sys.exit(f"xlsx not found: {xlsx}")
    out.mkdir(parents=True, exist_ok=True)
    phases,title,meta=read_plan(xlsx)
    n=sum(len(p["tasks"]) for p in phases)
    (out/"tracker.html").write_text(build_tracker(phases,title,meta),encoding="utf-8")
    (out/"gantt.html").write_text(build_gantt(phases,title,meta),encoding="utf-8")
    print(f"✓ Read {n} tasks across {len([p for p in phases if p['tasks']])} phases from {xlsx.name}")
    print(f"✓ Wrote {out/'tracker.html'}")
    print(f"✓ Wrote {out/'gantt.html'}")
    # public page
    repo=xlsx.resolve().parent.parent
    pages=repo/"pages"; tpl=pages/"agents.html"
    if pages.exists() and tpl.exists():
        (pages/"project-plan.html").write_text(build_public_plan(phases,meta,tpl),encoding="utf-8")
        print(f"✓ Wrote {pages/'project-plan.html'}  (public, chrome cloned from agents.html)")
    else:
        print(f"⚠ skipped public project-plan.html — template not found: {tpl}")
    print("\nNext: git add internal/ pages/project-plan.html && git commit -m 'plan: regen views' && git push")

if __name__=="__main__":
    main()
